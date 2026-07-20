from __future__ import annotations

import json
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from ashare_f10.export.bundle import build_exports


def fixture_combined() -> dict:
    return {
        "metadata": {
            "schema_version": "1.0.0",
            "security": {
                "code": "688521",
                "exchange": "SH",
                "secucode": "688521.SH",
                "page_code": "SH688521",
                "market_id": 1,
                "market_id_code": "1.688521",
            },
            "completed_at_utc": "2026-07-20T12:15:00Z",
            "fixed_manifest_version": "1.0.0",
            "group_count": 4,
            "completed_group_count": 4,
            "failed_group_count": 0,
            "source": "integration fixture",
        },
        "groups": [
            {
                "group_id": "q-income",
                "theme": "财务报表与指标",
                "family": "RPT_F10_FINANCE_GINCOMEQC",
                "strategy": "fixture",
                "success": True,
                "records": [
                    {"REPORT_DATE": "2025-06-30", "REPORT_TYPE": "Q2", "TOTAL_OPERATE_INCOME": 20.0},
                    {"REPORT_DATE": "2025-09-30", "REPORT_TYPE": "Q3", "TOTAL_OPERATE_INCOME": 30.0},
                    {"REPORT_DATE": "2025-12-31", "REPORT_TYPE": "Q4", "TOTAL_OPERATE_INCOME": 40.0},
                    {"REPORT_DATE": "2026-03-31", "REPORT_TYPE": "Q1", "TOTAL_OPERATE_INCOME": 10.0},
                ],
                "payloads": [],
                "requests": [],
            },
            {
                "group_id": "balance",
                "theme": "财务报表与指标",
                "family": "RPT_F10_FINANCE_GBALANCE",
                "strategy": "fixture",
                "success": True,
                "records": [
                    {"REPORT_DATE": "2026-03-31", "REPORT_TYPE": "Q1", "CIP": 25.0, "TOTAL_ASSETS": 100.0}
                ],
                "payloads": [],
                "requests": [],
            },
            {
                "group_id": "dupont-derived",
                "theme": "财务报表与指标",
                "family": "RPT_F10_FINANCE_DUPONT",
                "strategy": "fixture",
                "success": True,
                "records": [
                    {"REPORT_DATE": "2026-03-31", "REPORT_TYPE": "Q1", "CIP": 999.0, "TOTAL_ASSETS": 1000.0}
                ],
                "payloads": [],
                "requests": [],
            },
            {
                "group_id": "company",
                "theme": "公司资料",
                "family": "RPT_F10_ORG_BASICINFO",
                "strategy": "fixture",
                "success": True,
                "records": [{"SECURITY_CODE": "688521", "SECURITY_NAME_ABBR": "芯原股份"}],
                "payloads": [],
                "requests": [],
            },
        ],
    }


def test_exports_search_ttm_formula_and_downloads(tmp_path: Path, monkeypatch):
    data_dir = tmp_path / "data"
    run_dir = data_dir / "688521" / "integration"
    run_dir.mkdir(parents=True)
    combined = fixture_combined()
    (run_dir / "combined.json").write_text(json.dumps(combined, ensure_ascii=False), encoding="utf-8")
    artifacts = build_exports(combined, run_dir)
    (data_dir / "688521" / "latest.json").write_text(
        json.dumps({"job_id": "integration", "output_dir": str(run_dir), "artifacts": artifacts}),
        encoding="utf-8",
    )

    for key in ("json", "excel", "parquet", "duckdb"):
        assert Path(artifacts[key]).is_file()
    exported = json.loads(Path(artifacts["json"]).read_text(encoding="utf-8"))
    assert exported["metadata"]["security"]["code"] == "688521"
    assert exported["metadata"]["security"]["security_code"] == "688521"
    workbook = load_workbook(artifacts["excel"], read_only=True)
    assert "资产负债表" in workbook.sheetnames
    assert "单季度利润表" in workbook.sheetnames
    workbook.close()

    from ashare_f10.api import app as app_module

    monkeypatch.setattr(app_module.manager.settings, "data_dir", data_dir)
    monkeypatch.setattr(app_module.manager, "db_path", data_dir / "jobs.sqlite3")
    app_module.manager._init_db()
    client = TestClient(app_module.app)

    assert client.get("/api/health").status_code == 200
    overview = client.get("/api/stocks/688521/latest")
    assert overview.status_code == 200
    assert overview.json()["overview"]["fact_count"] > 0

    search = client.get("/api/stocks/688521/search", params={"q": "在建工程"})
    assert search.status_code == 200
    assert search.json()[0]["field_key"] == "CIP"

    ttm = client.post(
        "/api/stocks/688521/ttm",
        json={"field": "TOTAL_OPERATE_INCOME", "end_period": "2026-03-31"},
    )
    assert ttm.status_code == 200
    assert ttm.json()["value"] == 100.0
    assert ttm.json()["unit"] == "元"

    formula = client.post(
        "/api/stocks/688521/formula",
        json={"formula": 'F("在建工程") / F("资产总计")', "end_period": "2026-03-31"},
    )
    assert formula.status_code == 200
    assert formula.json()["value"] == 0.25
    assert [item["family"] for item in formula.json()["trace"]] == [
        "RPT_F10_FINANCE_GBALANCE",
        "RPT_F10_FINANCE_GBALANCE",
    ]

    for kind in ("json", "xlsx", "parquet", "db"):
        assert client.get(f"/api/stocks/688521/download/{kind}").status_code == 200
    assert client.get("/").status_code == 200
