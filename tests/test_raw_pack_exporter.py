from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from ashare_f10.raw_sources.models import (
    RawPackRun,
    SecurityEntity,
    SourceDocument,
    stable_document_id,
    utc_now,
)
from ashare_f10.raw_sources.raw_pack_exporter import export_raw_pack, validate_raw_pack


def document(status="FACT_DIRECT", suffix="1"):
    common = {
        "document_id": stable_document_id("688521", suffix),
        "security_code": "688521",
        "pack_id": "P0_STATUTORY_CORE",
        "source_id": "SRC001",
        "source_tier": "T0_STATUTORY",
        "source_organization": "上海证券交易所",
        "source_domain": "sse.com.cn",
        "source_url": f"https://www.sse.com.cn/{suffix}",
        "document_title": f"document-{suffix}",
        "status": status,
        "access_status": "DOWNLOAD_OK" if status == "FACT_DIRECT" else "NO_EXACT_HIT",
    }
    if status == "NO_MATCH":
        common.update(query="芯原", search_scope="fixture")
    return SourceDocument.model_validate(common)


def test_export_raw_pack_creates_index_and_quality(tmp_path):
    run = RawPackRun(
        run_id="test-run",
        security=SecurityEntity(
            security_code="688521",
            secucode="688521.SH",
            security_name_abbr="芯原股份",
            company_full_name_cn="芯原微电子（上海）股份有限公司",
            listed_market="SSE",
        ),
        started_at_utc=utc_now(),
        completed_at_utc=utc_now(),
        documents=[document(), document("NO_MATCH", "2")],
        output_dir=str(tmp_path),
    )
    artifacts = export_raw_pack(run, tmp_path)
    assert Path(artifacts.source_documents_jsonl).exists()
    assert Path(artifacts.excel_index).exists()
    quality = json.loads(Path(artifacts.quality_report_json).read_text())
    assert quality["status"] == "PASS"
    assert quality["document_count"] == 2
    workbook = load_workbook(artifacts.excel_index, read_only=True)
    assert {"来源文档索引", "质量摘要"}.issubset(workbook.sheetnames)


def test_duplicate_document_id_is_rejected_by_validator(tmp_path):
    path = tmp_path / "source_index"
    path.mkdir(parents=True)
    row = document().model_dump(mode="json")
    (path / "source_documents.jsonl").write_text(
        json.dumps(row, ensure_ascii=False) + "\n" + json.dumps(row, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    (tmp_path / "metadata").mkdir()
    (tmp_path / "metadata" / "raw_pack_run.json").write_text("{}")
    result = validate_raw_pack(tmp_path)
    assert result["status"] == "FAIL"
    assert "duplicate document_id" in result["failures"]
