from __future__ import annotations

import json
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import load_workbook

from ashare_f10.research_pack import run_research_pack
from ashare_f10.validation.models import OfficialFact

FACT_COLUMNS = [
    "security_code",
    "theme",
    "family",
    "dataset",
    "record_key",
    "report_date",
    "event_date",
    "period_type",
    "data_semantics",
    "field_key",
    "field_name_cn",
    "field_category",
    "value_text",
    "value_num",
    "unit",
    "source_url",
    "source_status",
]


def _fact(
    field_key: str,
    value: float,
    *,
    family: str,
    name: str,
    unit: str = "元",
    report_date: str = "2025-12-31",
) -> dict:
    semantics = "point_in_time" if "BALANCE" in family or "SHARE" in family else "flow"
    return {
        "security_code": "688521",
        "theme": "财务分析",
        "family": family,
        "dataset": family,
        "record_key": f"688521|{family}|{report_date}",
        "report_date": report_date,
        "event_date": report_date,
        "period_type": "FY",
        "data_semantics": semantics,
        "field_key": field_key,
        "field_name_cn": name,
        "field_category": "PAGE_DISPLAY_FIELD",
        "value_text": str(value),
        "value_num": value,
        "unit": unit,
        "source_url": "https://eastmoney.invalid/api",
        "source_status": "FACT_DIRECT",
    }


def _build_run_dir(path: Path, *, official_available_at: str = "2026-03-20") -> Path:
    normalized = path / "normalized"
    cross_validation = path / "cross_validation"
    normalized.mkdir(parents=True)
    cross_validation.mkdir(parents=True)
    facts = pd.DataFrame(
        [
            _fact(
                "OPERATE_INCOME",
                1_000_000_000.0,
                family="RPT_F10_FINANCE_GINCOME",
                name="营业收入",
            ),
            _fact(
                "PARENT_NETPROFIT",
                100_000_000.0,
                family="RPT_F10_FINANCE_GINCOME",
                name="归母净利润",
            ),
            _fact(
                "DEDUCT_PARENT_NETPROFIT",
                80_000_000.0,
                family="RPT_F10_FINANCE_MAINFINADATA",
                name="扣非归母净利润",
            ),
            _fact(
                "NETCASH_OPERATE",
                150_000_000.0,
                family="RPT_F10_FINANCE_GCASHFLOW",
                name="经营活动现金流量净额",
            ),
            _fact(
                "CONSTRUCT_LONG_ASSET",
                30_000_000.0,
                family="RPT_F10_FINANCE_GCASHFLOW",
                name="购建长期资产支付现金",
            ),
            _fact(
                "RESEARCH_EXPENSE",
                20_000_000.0,
                family="RPT_F10_FINANCE_GINCOME",
                name="研发费用",
            ),
            _fact(
                "TOTAL_SHARES",
                1_000_000_000.0,
                family="RPT_F10_SHARE_STRUCTURE",
                name="总股本",
                unit="股",
            ),
        ],
        columns=FACT_COLUMNS,
    )
    facts_path = normalized / "facts.parquet"
    facts.to_parquet(facts_path, index=False)
    connection = duckdb.connect(str(normalized / "f10.duckdb"))
    try:
        connection.execute("CREATE TABLE facts AS SELECT * FROM read_parquet(?)", [str(facts_path)])
    finally:
        connection.close()

    official = OfficialFact(
        security_code="688521",
        report_date="2025-12-31",
        statement_type="income_statement",
        scope="consolidated",
        field_key="OPERATE_INCOME",
        field_name_report="营业收入",
        value=1_000_000_000.0,
        unit="元",
        normalized_unit="元",
        source_document="芯原股份2025年年度报告",
        source_url="https://sse.invalid/annual.pdf",
        source_page=80,
        source_row="营业收入 1,000,000,000.00",
        extraction_method="PDF_TABLE",
        precision_tolerance=0.01,
        confidence="high",
        document_id="annual-2025",
        available_at=official_available_at,
    )
    pd.DataFrame([official.to_dict()]).to_parquet(
        cross_validation / "official_direct_facts.parquet",
        index=False,
    )
    (path / "artifacts.json").write_text(
        json.dumps({"duckdb": str(normalized / "f10.duckdb")}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def test_research_pack_exports_all_formats_and_passes_quality(tmp_path: Path) -> None:
    run_dir = _build_run_dir(tmp_path / "run")
    result = run_research_pack("688521", run_dir, as_of_date="2026-03-31")
    assert result["status"] == "COMPLETED"
    assert result["cache_hit"] is False
    artifacts = {key: Path(value) for key, value in result["artifacts"].items()}
    for key in (
        "manifest_json",
        "summary_json",
        "package_json",
        "package_excel",
        "package_duckdb",
        "quality_json",
        "checkpoint_json",
    ):
        assert artifacts[key].exists() and artifacts[key].stat().st_size > 0

    quality = json.loads(artifacts["quality_json"].read_text(encoding="utf-8"))
    assert quality["status"] == "PASS"
    assert quality["failures"] == []

    workbook = load_workbook(artifacts["package_excel"], read_only=True)
    try:
        assert {"Summary", "CanonicalFacts", "FactLineage", "ProfitQuality"}.issubset(
            set(workbook.sheetnames)
        )
    finally:
        workbook.close()

    connection = duckdb.connect(str(artifacts["package_duckdb"]), read_only=True)
    try:
        tables = {row[0] for row in connection.execute("SHOW TABLES").fetchall()}
        assert {
            "source_facts",
            "canonical_observations",
            "fact_lineage",
            "evidence_nodes",
            "evidence_edges",
            "profit_quality",
        }.issubset(tables)
        assert connection.execute("SELECT count(*) FROM resolved_canonical_facts").fetchone()[0] > 0
    finally:
        connection.close()

    summary = json.loads(artifacts["summary_json"].read_text(encoding="utf-8"))
    assert summary["mapping_coverage"]["source_fact_count"] >= 7
    assert summary["evidence_quality"]["status"] == "PASS"
    assert summary["section_summary"]["calculated_fact_count"] > 0


def test_second_run_reuses_completed_pack_without_recomputing(tmp_path: Path) -> None:
    run_dir = _build_run_dir(tmp_path / "run")
    first = run_research_pack("688521", run_dir, as_of_date="2026-03-31")
    json_path = Path(first["artifacts"]["package_json"])
    first_mtime = json_path.stat().st_mtime_ns
    second = run_research_pack("688521", run_dir, as_of_date="2026-03-31")
    assert second["status"] == "COMPLETED"
    assert second["cache_hit"] is True
    assert json_path.stat().st_mtime_ns == first_mtime


def test_as_of_date_excludes_future_official_fact_but_keeps_source_history(tmp_path: Path) -> None:
    run_dir = _build_run_dir(tmp_path / "run", official_available_at="2026-04-01")
    result = run_research_pack("688521", run_dir, as_of_date="2026-03-31")
    output_dir = Path(result["artifacts"]["output_dir"])
    source_facts = pd.read_parquet(output_dir / "tables" / "source_facts.parquet")
    revenue = source_facts[source_facts["metric_id"] == "financial.revenue"]
    assert set(revenue["source_name"]) == {"EASTMONEY"}
    canonical = pd.read_parquet(output_dir / "tables" / "canonical_observations.parquet")
    revenue_observation = canonical[canonical["metric_id"] == "financial.revenue"].iloc[0]
    assert revenue_observation["status"] == "SINGLE_SOURCE"


def test_checkpoint_contains_recovery_stages_and_artifacts_registration(tmp_path: Path) -> None:
    run_dir = _build_run_dir(tmp_path / "run")
    result = run_research_pack("688521", run_dir, as_of_date="2026-03-31")
    checkpoint = json.loads(
        Path(result["artifacts"]["checkpoint_json"]).read_text(encoding="utf-8")
    )
    assert checkpoint["status"] == "COMPLETED"
    assert checkpoint["completed_stages"] == [
        "LOAD_INPUTS",
        "MAP_CANONICAL",
        "EXTRACT_SECTIONS",
        "BUILD_EVIDENCE",
        "EXPORT",
        "VALIDATE",
    ]
    assert checkpoint["retry_queue"] == []
    artifacts = json.loads((run_dir / "artifacts.json").read_text(encoding="utf-8"))
    assert Path(artifacts["research_pack_excel"]).exists()
    assert Path(artifacts["research_pack_quality"]).exists()
