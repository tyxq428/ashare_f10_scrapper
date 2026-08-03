from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ashare_f10.export.excel_exporter import export_excel
from ashare_f10.export.excel_sanitizer import sanitize_excel_payload_in_place


def _combined(text: str) -> dict:
    return {
        "metadata": {
            "security": {"code": "000403", "secucode": "000403.SZ"},
            "completed_at_utc": "2026-07-30T15:05:00Z",
            "group_count": 1,
            "completed_group_count": 1,
            "failed_group_count": 0,
            "source": "fixture",
        },
        "groups": [
            {
                "group_id": "illegal-character-regression",
                "record_count": 1,
                "family": "TEST_NON_FINANCIAL",
                "strategy": "fixture",
                "theme": "回归测试",
                "records": [{"TEXT": text}],
            }
        ],
    }


def test_sanitizer_removes_only_excel_forbidden_control_characters() -> None:
    payload = _combined("A\x00B\x02C\x19D\tline\nnext\r")

    report = sanitize_excel_payload_in_place(payload)

    assert report.affected_strings == 1
    assert report.removed_characters == 3
    assert report.removed_codepoints == {"U+0000": 1, "U+0002": 1, "U+0019": 1}
    assert payload["groups"][0]["records"][0]["TEXT"] == "ABCD\tline\nnext\r"


def test_sanitized_payload_can_be_written_to_excel(tmp_path: Path) -> None:
    payload = _combined("判决内容\x00和模块\x02以及公告\x19")
    sanitize_excel_payload_in_place(payload)

    path = export_excel(payload, tmp_path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    values = [
        cell.value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]
    assert "判决内容和模块以及公告" in values
