from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from ashare_f10.export.excel_exporter import export_excel
from ashare_f10.fetch.manifest import load_field_mapping
from ashare_f10.models import GroupResult


def _cashflow_record(date: str, name: str, main: float, note: float | None) -> dict:
    # Preserve the API's field order: primary statement first, then the separate
    # supplementary reconciliation section, followed by growth fields.
    return {
        "SECUCODE": "300308.SZ",
        "SECURITY_CODE": "300308",
        "SECURITY_NAME_ABBR": "中际旭创",
        "REPORT_DATE": f"{date} 00:00:00",
        "REPORT_DATE_NAME": name,
        "NETCASH_OPERATE": main,
        "CONSTRUCT_LONG_ASSET": main / 2,
        "NETPROFIT": main / 3,
        "NETCASH_OPERATENOTE": note,
        "END_CASH": note,
        "CCE_ADDNOTE": note,
        "NETCASH_OPERATE_YOY": 12.3,
    }


def _combined(records: list[dict]) -> dict:
    return {
        "metadata": {
            "security": {"code": "300308", "secucode": "300308.SZ"},
            "completed_at_utc": "2026-07-20T00:00:00Z",
            "group_count": 1,
            "completed_group_count": 1,
            "failed_group_count": 0,
            "source": "Eastmoney live APIs",
        },
        "groups": [
            {
                "group_id": "cashflow-test",
                "theme": "财务报表与指标",
                "family": "RPT_F10_FINANCE_GCASHFLOW",
                "strategy": "finance_all_periods",
                "success": True,
                "used_fallback": False,
                "record_count": len(records),
                "records": records,
                "payloads": [],
                "requests": [],
                "errors": [],
                "started_at_utc": "2026-07-20T00:00:00Z",
                "completed_at_utc": "2026-07-20T00:00:01Z",
            }
        ],
    }


def _find_key_row(ws, key: str) -> int | None:
    for row in range(3, ws.max_row + 1):
        if ws.cell(row, 2).value == key:
            return row
    return None


def test_cashflow_primary_and_supplement_are_separated(tmp_path: Path) -> None:
    records = [
        _cashflow_record("2026-03-31", "2026一季报", 3_367_573_676.62, None),
        _cashflow_record("2025-12-31", "2025年报", 10_896_126_160.03, 10_896_126_160.03),
        _cashflow_record("2025-09-30", "2025三季报", 5_454_821_227.43, None),
        _cashflow_record("2025-06-30", "2025中报", 3_218_463_287.30, 3_218_463_287.30),
    ]
    path = export_excel(_combined(records), tmp_path)
    book = load_workbook(path, data_only=True)

    assert "现金流量表" in book.sheetnames
    assert "现金流补充资料" in book.sheetnames
    assert "现金流量表同比" in book.sheetnames

    primary = book["现金流量表"]
    supplement = book["现金流补充资料"]

    main_row = _find_key_row(primary, "NETCASH_OPERATE")
    note_row = _find_key_row(supplement, "NETCASH_OPERATENOTE")
    assert main_row is not None
    assert note_row is not None
    assert _find_key_row(primary, "NETCASH_OPERATENOTE") is None

    assert primary.cell(main_row, 1).value == "经营活动产生的现金流量净额"
    assert primary.cell(main_row, 4).value == pytest.approx(3_367_573_676.62)
    assert primary.cell(main_row, 6).value == pytest.approx(5_454_821_227.43)

    assert supplement.cell(note_row, 1).value == "经营活动产生的现金流量净额（补充资料）"
    assert supplement.cell(note_row, 4).value is None
    assert supplement.cell(note_row, 5).value == pytest.approx(10_896_126_160.03)
    assert supplement.cell(note_row, 6).value is None
    assert supplement.cell(note_row, 7).value == pytest.approx(3_218_463_287.30)


def test_cashflow_mapping_distinguishes_primary_and_note() -> None:
    mapping = load_field_mapping()["global"]
    assert mapping["NETCASH_OPERATE"]["label"] == "经营活动产生的现金流量净额"
    assert mapping["NETCASH_OPERATENOTE"]["label"] == "经营活动产生的现金流量净额（补充资料）"
    assert mapping["NETCASH_OPERATE"]["unit"] == "元"


def test_group_result_rejects_record_count_mismatch() -> None:
    with pytest.raises(ValidationError, match="record_count=64"):
        GroupResult(
            group_id="bad-cache",
            theme="财务报表与指标",
            family="RPT_F10_FINANCE_GCASHFLOW",
            strategy="finance_all_periods",
            success=True,
            record_count=64,
            records=[],
            started_at_utc="2026-07-20T00:00:00Z",
            completed_at_utc="2026-07-20T00:00:01Z",
        )
