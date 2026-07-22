from __future__ import annotations

from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]


def replace_once(path: Path, old: str, new: str, label: str) -> None:
    text = path.read_text(encoding="utf-8")
    if old not in text:
        if new in text:
            return
        raise SystemExit(f"{label} marker not found in {path}")
    path.write_text(text.replace(old, new, 1), encoding="utf-8")


# Pandas does not guarantee row-major writes when using XlsxWriter.  XlsxWriter's
# constant-memory mode, however, discards cells written after the next row starts.
# The combination silently produced workbooks where headers had many columns but
# every data row only retained column A.  These workbooks are intentionally summary
# products, so normal worksheet buffering is both safe and correct.
exporter = ROOT / "src/ashare_f10/cross_validation/exporter.py"
replace_once(
    exporter,
    '        engine_kwargs={"options": {"constant_memory": True, "strings_to_urls": False}},\n',
    '        engine_kwargs={"options": {"strings_to_urls": False}},\n',
    "remove unsupported pandas/xlsxwriter constant-memory combination",
)

verify = ROOT / "scripts/verify_full_cross_validation.py"
replace_once(
    verify,
    "import pandas as pd\nfrom fastapi.testclient import TestClient\n",
    "import pandas as pd\nfrom fastapi.testclient import TestClient\nfrom openpyxl import load_workbook\n",
    "import openpyxl verifier",
)
helper = '''\n\ndef assert_tabular_xlsx(path: Path) -> None:\n    \"\"\"Fail when a multi-column worksheet silently collapses data into column A.\"\"\"\n\n    workbook = load_workbook(path, read_only=True, data_only=True)\n    checked_sheets = 0\n    try:\n        for sheet in workbook.worksheets:\n            if sheet.max_row <= 1 or sheet.max_column <= 1:\n                continue\n            header = next(\n                sheet.iter_rows(min_row=1, max_row=1, values_only=True),\n                (),\n            )\n            if sum(value not in (None, \"\") for value in header) < 2:\n                continue\n            rows = sheet.iter_rows(\n                min_row=2,\n                max_row=min(sheet.max_row, 500),\n                max_col=sheet.max_column,\n                values_only=True,\n            )\n            found_multicolumn_row = any(\n                sum(value not in (None, \"\") for value in row) >= 2\n                for row in rows\n            )\n            if not found_multicolumn_row:\n                raise AssertionError(\n                    f\"{path.name}/{sheet.title}: multi-column data collapsed into column A\"\n                )\n            checked_sheets += 1\n    finally:\n        workbook.close()\n    if checked_sheets == 0:\n        raise AssertionError(f\"No populated multi-column worksheet found in {path}\")\n'''
replace_once(
    verify,
    "\ndef main(run_dir: Path) -> None:\n",
    helper + "\n\ndef main(run_dir: Path) -> None:\n",
    "add workbook matrix validator",
)
replace_once(
    verify,
    '''    for key in ("eastmoney_excel", "official_excel", "comparison_excel"):\n        with zipfile.ZipFile(paths[key]) as archive:\n            assert archive.testzip() is None\n            assert "xl/workbook.xml" in archive.namelist()\n''',
    '''    for key in ("eastmoney_excel", "official_excel", "comparison_excel"):\n        with zipfile.ZipFile(paths[key]) as archive:\n            assert archive.testzip() is None\n            assert "xl/workbook.xml" in archive.namelist()\n    # The themed Eastmoney workbook uses a different openpyxl writer and is covered\n    # by its own regressions.  The two pandas/XlsxWriter packages must additionally\n    # prove that data rows occupy their declared columns rather than only column A.\n    assert_tabular_xlsx(paths["official_excel"])\n    assert_tabular_xlsx(paths["comparison_excel"])\n''',
    "invoke workbook matrix validator",
)

web = ROOT / "src/ashare_f10/web/cross-validation.js"
replace_once(
    web,
    '''    const metrics = [\n      ["字段分类覆盖率", `${((summary.classification_coverage || 0) * 100).toFixed(2)}%`],\n      ["东方财富事实", summary.eastmoney_fact_count || 0],\n      ["官方来源", sourceStatus.source || "—"],\n      ["官方事实", summary.official_fact_count || 0],\n      ["可比记录", sourceAvailable ? (summary.comparable_count || 0) : "—"],\n      ["已匹配", sourceAvailable ? (summary.matched_count || 0) : "—"],\n      ["真正冲突", sourceAvailable ? (summary.true_conflict_count || 0) : "—"],\n      ["验收状态", summary.acceptance_status || "—"],\n    ];\n''',
    '''    const statusCounts = summary.status_counts || {};\n    const metrics = [\n      ["字段分类覆盖率", `${((summary.classification_coverage || 0) * 100).toFixed(2)}%`],\n      ["东方财富事实", summary.eastmoney_fact_count || 0],\n      ["官方来源", sourceStatus.source || "—"],\n      ["官方事实", summary.official_fact_count || 0],\n      ["理论可比记录", sourceAvailable ? (summary.comparable_count || 0) : "—"],\n      ["已形成双源匹配", sourceAvailable ? (summary.matched_count || 0) : "—"],\n      ["待官方提取", sourceAvailable ? (statusCounts.MISSING_OFFICIAL || 0) : "—"],\n      ["官方有而东方财富缺失", sourceAvailable ? (statusCounts.MISSING_EASTMONEY || 0) : "—"],\n      ["真正冲突", sourceAvailable ? (summary.true_conflict_count || 0) : "—"],\n      ["验收状态", summary.acceptance_status || "—"],\n    ];\n''',
    "clarify comparable and matched metrics",
)

index = ROOT / "src/ashare_f10/web/index.html"
replace_once(
    index,
    '        <label>状态 <select id="validationStatus"><option value="">全部</option><option>EXACT_MATCH</option><option>WITHIN_ROUNDING</option><option>DERIVED_MATCH</option><option>MISMATCH</option><option>MISSING_OFFICIAL</option><option>OFFICIAL_SOURCE_UNAVAILABLE</option><option>MISSING_EASTMONEY</option><option>NOT_IN_OFFICIAL_SCOPE</option><option>SOURCE_SPECIFIC</option><option>FUTURE_FREE_SOURCE_REQUIRED</option></select></label>\n',
    '        <label>状态 <select id="validationStatus"><option value="">全部</option><option value="EXACT_MATCH">完全一致（EXACT_MATCH）</option><option value="WITHIN_ROUNDING">披露精度内一致（WITHIN_ROUNDING）</option><option value="DERIVED_MATCH">官方派生一致（DERIVED_MATCH）</option><option value="TEXT_MATCH_NORMALIZED">文本标准化一致（TEXT_MATCH_NORMALIZED）</option><option value="SET_MATCH">集合一致（SET_MATCH）</option><option value="MISMATCH">数值冲突（MISMATCH）</option><option value="VERSION_CONFLICT">报告版本冲突（VERSION_CONFLICT）</option><option value="SCOPE_CONFLICT">合并/母公司口径冲突（SCOPE_CONFLICT）</option><option value="PERIOD_CONFLICT">期间口径冲突（PERIOD_CONFLICT）</option><option value="UNIT_CONFLICT">单位冲突（UNIT_CONFLICT）</option><option value="MISSING_OFFICIAL">待官方提取（MISSING_OFFICIAL）</option><option value="OFFICIAL_PERIOD_NOT_LOADED">官方报告期未加载（OFFICIAL_PERIOD_NOT_LOADED）</option><option value="OFFICIAL_SOURCE_UNAVAILABLE">官方来源未接入（OFFICIAL_SOURCE_UNAVAILABLE）</option><option value="MISSING_EASTMONEY">东方财富缺失（MISSING_EASTMONEY）</option><option value="NOT_IN_OFFICIAL_SCOPE">不在定期报告范围（NOT_IN_OFFICIAL_SCOPE）</option><option value="SOURCE_SPECIFIC">东方财富专有（SOURCE_SPECIFIC）</option><option value="FUTURE_FREE_SOURCE_REQUIRED">待其他免费来源（FUTURE_FREE_SOURCE_REQUIRED）</option><option value="UNRESOLVED">未解决（UNRESOLVED）</option></select></label>\n',
    "expand and translate status filter",
)
replace_once(
    index,
    '        <label>验证模式 <select id="validationMode"><option value="">全部</option><option>OFFICIAL_DIRECT</option><option>OFFICIAL_DERIVED</option><option>OFFICIAL_DOCUMENT_EVENT</option><option>OFFICIAL_METADATA</option><option>NOT_IN_PERIODIC_REPORT_SCOPE</option><option>EASTMONEY_SOURCE_SPECIFIC</option><option>FUTURE_FREE_SOURCE_REQUIRED</option></select></label>\n',
    '        <label>验证模式 <select id="validationMode"><option value="">全部</option><option value="OFFICIAL_DIRECT">官方直接披露（OFFICIAL_DIRECT）</option><option value="OFFICIAL_DERIVED">由官方事实计算（OFFICIAL_DERIVED）</option><option value="OFFICIAL_DOCUMENT_EVENT">官方公告事件（OFFICIAL_DOCUMENT_EVENT）</option><option value="OFFICIAL_METADATA">官方元数据（OFFICIAL_METADATA）</option><option value="NOT_IN_PERIODIC_REPORT_SCOPE">不在定期报告范围（NOT_IN_PERIODIC_REPORT_SCOPE）</option><option value="EASTMONEY_SOURCE_SPECIFIC">东方财富专有口径（EASTMONEY_SOURCE_SPECIFIC）</option><option value="FUTURE_FREE_SOURCE_REQUIRED">待其他免费来源（FUTURE_FREE_SOURCE_REQUIRED）</option></select></label>\n',
    "translate validation-mode filter",
)
