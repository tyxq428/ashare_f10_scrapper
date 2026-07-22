from __future__ import annotations

import json
import sys
import urllib.parse
import zipfile
from pathlib import Path

import duckdb
import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook


def require(path: Path, minimum_size: int = 1) -> Path:
    if not path.exists() or not path.is_file() or path.stat().st_size < minimum_size:
        raise AssertionError(f"Missing or empty artifact: {path}")
    return path


def assert_tabular_xlsx(path: Path) -> None:
    """Fail when a multi-column worksheet silently collapses data into column A."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    checked_sheets = 0
    try:
        for sheet in workbook.worksheets:
            if sheet.max_row <= 1 or sheet.max_column <= 1:
                continue
            header = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=True),
                (),
            )
            if sum(value not in (None, "") for value in header) < 2:
                continue
            rows = sheet.iter_rows(
                min_row=2,
                max_row=min(sheet.max_row, 500),
                max_col=sheet.max_column,
                values_only=True,
            )
            found_multicolumn_row = any(sum(value not in (None, "") for value in row) >= 2 for row in rows)
            if not found_multicolumn_row:
                raise AssertionError(f"{path.name}/{sheet.title}: multi-column data collapsed into column A")
            checked_sheets += 1
    finally:
        workbook.close()
    if checked_sheets == 0:
        raise AssertionError(f"No populated multi-column worksheet found in {path}")


def main(run_dir: Path) -> None:
    cross_dir = run_dir / "cross_validation"
    summary_path = require(cross_dir / "cross_validation_summary.json")
    summary = json.loads(summary_path.read_text(encoding="utf-8"))

    assert summary["paid_sources_used"] is False
    assert float(summary["classification_coverage"]) == 1.0
    assert int(summary["classified_field_contexts"]) == int(summary["unique_field_contexts"])
    assert int(summary["true_conflict_count"]) == 0
    assert summary["acceptance_status"] in {"PASS", "PASS_WITH_COVERAGE_GAPS"}
    assert int(summary["eastmoney_fact_count"]) > 1000
    assert int(summary["official_fact_count"]) >= 20

    artifacts = summary["artifacts"]
    required_keys = {
        "eastmoney_json",
        "eastmoney_excel",
        "eastmoney_parquet",
        "eastmoney_duckdb",
        "official_json",
        "official_excel",
        "official_parquet",
        "official_duckdb",
        "comparison_json",
        "comparison_excel",
        "comparison_parquet",
        "comparison_duckdb",
        "summary_json",
        "evidence_zip",
    }
    assert required_keys.issubset(artifacts)
    paths = {key: require(Path(artifacts[key]), 20) for key in required_keys}

    east = pd.read_parquet(paths["eastmoney_parquet"])
    official = pd.read_parquet(paths["official_parquet"])
    comparison = pd.read_parquet(paths["comparison_parquet"])
    assert len(east) == int(summary["eastmoney_fact_count"])
    assert len(official) == int(summary["official_fact_count"])
    assert len(comparison) == int(summary["comparison_count"])
    assert comparison["validation_mode"].notna().all()
    assert "EXACT_MATCH" in set(comparison["status"])
    assert set(comparison["status"]) & {
        "NOT_IN_OFFICIAL_SCOPE",
        "SOURCE_SPECIFIC",
        "FUTURE_FREE_SOURCE_REQUIRED",
    }
    assert (
        not comparison["status"]
        .isin(["MISMATCH", "VERSION_CONFLICT", "SCOPE_CONFLICT", "PERIOD_CONFLICT", "UNIT_CONFLICT"])
        .any()
    )

    connection = duckdb.connect(str(paths["comparison_duckdb"]), read_only=True)
    try:
        tables = {row[0] for row in connection.execute("SHOW TABLES").fetchall()}
        expected_tables = {
            "eastmoney_facts",
            "official_facts",
            "field_validation_registry",
            "reconciliation",
            "logic_checks",
            "ttm_checks",
            "documents",
            "report_period_lifecycle",
        }
        assert expected_tables.issubset(tables)
        assert connection.execute("SELECT count(*) FROM true_conflicts").fetchone()[0] == 0
        assert (
            connection.execute(
                "SELECT count(*) FROM field_validation_registry WHERE validation_mode IS NULL"
            ).fetchone()[0]
            == 0
        )
        sample = connection.execute(
            "SELECT comparison_key FROM reconciliation WHERE status='EXACT_MATCH' LIMIT 1"
        ).fetchone()
        assert sample
        comparison_key = str(sample[0])
    finally:
        connection.close()

    for key in ("eastmoney_excel", "official_excel", "comparison_excel"):
        with zipfile.ZipFile(paths[key]) as archive:
            assert archive.testzip() is None
            assert "xl/workbook.xml" in archive.namelist()
    # The themed Eastmoney workbook uses a different openpyxl writer and is covered
    # by its own regressions.  The two pandas/XlsxWriter packages must additionally
    # prove that data rows occupy their declared columns rather than only column A.
    assert_tabular_xlsx(paths["official_excel"])
    assert_tabular_xlsx(paths["comparison_excel"])
    with zipfile.ZipFile(paths["evidence_zip"]) as archive:
        assert archive.testzip() is None
        names = archive.namelist()
        assert "cross_validation_summary.json" in names
        assert any(name.startswith("source_documents/") and name.endswith(".pdf") for name in names)

    # Exercise the integrated FastAPI router without running a second network job.
    from ashare_f10.api.app import app
    from ashare_f10.cross_validation import api as cross_api

    task_id = "e2e-validation-task"
    cross_api._tasks[task_id] = {
        "task_id": task_id,
        "stock_code": str(summary["security_code"]),
        "status": "COMPLETED",
        "stage": "COMPLETED",
        "summary": summary,
        "artifacts": artifacts,
    }
    client = TestClient(app)
    assert client.get(f"/api/cross-validation/jobs/{task_id}").status_code == 200
    coverage = client.get(f"/api/cross-validation/jobs/{task_id}/coverage")
    assert coverage.status_code == 200
    assert coverage.json()["classification_coverage"] == 1.0
    rows = client.get(
        f"/api/cross-validation/jobs/{task_id}/comparison",
        params={"q": "资产", "limit": 20},
    )
    assert rows.status_code == 200 and rows.json()["total"] > 0
    encoded_key = urllib.parse.quote(comparison_key, safe="")
    evidence = client.get(f"/api/cross-validation/jobs/{task_id}/evidence/{encoded_key}")
    assert evidence.status_code == 200
    for kind in ("comparison_xlsx", "comparison_db", "official_xlsx", "eastmoney_xlsx", "evidence"):
        response = client.get(f"/api/cross-validation/jobs/{task_id}/download/{kind}")
        assert response.status_code == 200 and len(response.content) > 20

    report = {
        "status": "PASS",
        "security_code": summary["security_code"],
        "acceptance_status": summary["acceptance_status"],
        "classification_coverage": summary["classification_coverage"],
        "eastmoney_fact_count": len(east),
        "official_fact_count": len(official),
        "comparison_count": len(comparison),
        "true_conflict_count": summary["true_conflict_count"],
        "status_counts": summary["status_counts"],
        "mode_counts": summary["mode_counts"],
        "artifacts": artifacts,
    }
    output = run_dir / "full-cross-validation-e2e-summary.json"
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main(Path(sys.argv[1] if len(sys.argv) > 1 else "full-cross-validation-run"))
