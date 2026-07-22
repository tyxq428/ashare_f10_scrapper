from __future__ import annotations

import json
import shutil
import sys
import time
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
from openpyxl import load_workbook

from ashare_f10.research_pack import run_research_pack
from ashare_f10.research_pack.exporter import validate_research_pack
from ashare_f10.research_pack.models import ResearchPackArtifacts


def require(path: Path, minimum_size: int = 1) -> Path:
    if not path.exists() or not path.is_file() or path.stat().st_size < minimum_size:
        raise AssertionError(f"Missing or empty artifact: {path}")
    return path


def stream_json_sanity(path: Path) -> None:
    """Check a potentially large JSON package without loading it all into memory."""

    depth = 0
    in_string = False
    escaped = False
    first_token = ""
    last_token = ""
    with path.open("r", encoding="utf-8") as handle:
        while chunk := handle.read(1024 * 1024):
            for character in chunk:
                if not character.isspace():
                    if not first_token:
                        first_token = character
                    last_token = character
                if in_string:
                    if escaped:
                        escaped = False
                    elif character == "\\":
                        escaped = True
                    elif character == '"':
                        in_string = False
                    continue
                if character == '"':
                    in_string = True
                elif character in "[{":
                    depth += 1
                elif character in "]}":
                    depth -= 1
                    if depth < 0:
                        raise AssertionError(f"Unbalanced JSON package: {path}")
    if in_string or depth != 0 or first_token != "{" or last_token != "}":
        raise AssertionError(f"Truncated or malformed JSON package: {path}")


def copied_artifacts(copy_dir: Path, security_code: str) -> ResearchPackArtifacts:
    return ResearchPackArtifacts(
        output_dir=copy_dir,
        manifest_json=copy_dir / "manifest.json",
        summary_json=copy_dir / "summary.json",
        package_json=copy_dir / "exports" / f"{security_code}_research_pack.json",
        package_excel=copy_dir / "exports" / f"{security_code}_research_pack.xlsx",
        package_duckdb=copy_dir / "exports" / f"{security_code}_research_pack.duckdb",
        quality_json=copy_dir / "quality" / "research_pack_quality.json",
        checkpoint_json=copy_dir / "checkpoint.json",
    )


def table_counts(database: Path) -> dict[str, int]:
    connection = duckdb.connect(str(database), read_only=True)
    try:
        tables = [row[0] for row in connection.execute("SHOW TABLES").fetchall()]
        counts = {
            table: int(connection.execute(f'SELECT count(*) FROM "{table}"').fetchone()[0])
            for table in tables
        }
        required = {
            "source_facts",
            "canonical_observations",
            "fact_lineage",
            "evidence_nodes",
            "evidence_edges",
            "profit_quality",
            "coverage_gaps",
        }
        if not required.issubset(tables):
            raise AssertionError(f"Research Pack DuckDB missing tables: {sorted(required - set(tables))}")
        duplicate_sources = connection.execute(
            "SELECT count(*) - count(DISTINCT source_fact_id) FROM source_facts"
        ).fetchone()[0]
        duplicate_observations = connection.execute(
            "SELECT count(*) - count(DISTINCT observation_id) FROM canonical_observations"
        ).fetchone()[0]
        missing_lineage = connection.execute(
            """
            SELECT count(*) FROM canonical_observations observation
            LEFT JOIN fact_lineage lineage USING (observation_id)
            WHERE lineage.observation_id IS NULL
            """
        ).fetchone()[0]
        dangling_edges = connection.execute(
            """
            SELECT count(*) FROM evidence_edges edge
            LEFT JOIN evidence_nodes source_node ON edge.from_node_id = source_node.node_id
            LEFT JOIN evidence_nodes target_node ON edge.to_node_id = target_node.node_id
            WHERE source_node.node_id IS NULL OR target_node.node_id IS NULL
            """
        ).fetchone()[0]
        if any((duplicate_sources, duplicate_observations, missing_lineage, dangling_edges)):
            raise AssertionError(
                "Research Pack integrity failure: "
                f"duplicate_sources={duplicate_sources}, "
                f"duplicate_observations={duplicate_observations}, "
                f"missing_lineage={missing_lineage}, dangling_edges={dangling_edges}"
            )
        return counts
    finally:
        connection.close()


def verify_cross_validation(run_dir: Path) -> dict[str, Any]:
    summary_path = run_dir / "cross_validation" / "cross_validation_summary.json"
    if not summary_path.exists():
        return {"available": False}
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    if str(summary.get("acceptance_status", "")).startswith("FAIL"):
        raise AssertionError(f"Cross validation failed: {summary.get('acceptance_status')}")
    if int(summary.get("true_conflict_count", 0)) != 0:
        raise AssertionError(f"Unexpected true conflicts: {summary.get('true_conflict_count')}")
    return {
        "available": True,
        "acceptance_status": summary.get("acceptance_status"),
        "comparison_coverage": summary.get("comparison_coverage"),
        "comparison_accuracy": summary.get("comparison_accuracy"),
        "true_conflict_count": summary.get("true_conflict_count"),
        "logic_check_counts": summary.get("logic_check_counts", {}),
        "ttm_check_counts": summary.get("ttm_check_counts", {}),
    }


def main(run_dir: Path, security_code: str) -> None:
    output_dir = run_dir / "research_pack_w08"
    started = time.perf_counter()
    first = run_research_pack(
        security_code,
        run_dir,
        output_dir,
        as_of_date="2026-07-22",
        force=True,
    )
    first_seconds = time.perf_counter() - started
    if first["status"] != "COMPLETED" or first["cache_hit"]:
        raise AssertionError(f"Unexpected first Research Pack result: {first}")

    artifacts = {key: Path(value) for key, value in first["artifacts"].items()}
    package_json = require(artifacts["package_json"], 20)
    package_excel = require(artifacts["package_excel"], 20)
    package_duckdb = require(artifacts["package_duckdb"], 20)
    quality_path = require(artifacts["quality_json"], 20)
    stream_json_sanity(package_json)
    quality = json.loads(quality_path.read_text(encoding="utf-8"))
    if quality.get("status") != "PASS":
        raise AssertionError(f"Research Pack quality failed: {quality.get('failures')}")

    workbook = load_workbook(package_excel, read_only=True, data_only=False)
    try:
        required_sheets = {"Summary", "SourceFacts", "CanonicalFacts", "FactLineage"}
        if not required_sheets.issubset(workbook.sheetnames):
            raise AssertionError(
                f"Research Pack workbook missing sheets: {sorted(required_sheets - set(workbook.sheetnames))}"
            )
    finally:
        workbook.close()
    counts = table_counts(package_duckdb)

    table_dir = output_dir / "tables"
    parquet_rows: dict[str, int] = {}
    for path in sorted(table_dir.glob("*.parquet")):
        parquet_rows[path.stem] = len(pd.read_parquet(path, columns=[]))
    if not parquet_rows:
        raise AssertionError("Research Pack did not export Parquet tables")

    package_mtime = package_json.stat().st_mtime_ns
    started = time.perf_counter()
    second = run_research_pack(
        security_code,
        run_dir,
        output_dir,
        as_of_date="2026-07-22",
    )
    second_seconds = time.perf_counter() - started
    if second["status"] != "COMPLETED" or not second["cache_hit"]:
        raise AssertionError(f"Research Pack cache reuse failed: {second}")
    if package_json.stat().st_mtime_ns != package_mtime:
        raise AssertionError("Cache-hit run rewrote the package JSON")

    portable_dir = run_dir / "research_pack_w08_portable_copy"
    if portable_dir.exists():
        shutil.rmtree(portable_dir)
    shutil.copytree(output_dir, portable_dir)
    portable_quality = validate_research_pack(copied_artifacts(portable_dir, security_code))
    if portable_quality["status"] != "PASS":
        raise AssertionError(f"Portable Research Pack validation failed: {portable_quality['failures']}")
    portable_counts = table_counts(
        portable_dir / "exports" / f"{security_code}_research_pack.duckdb"
    )
    if portable_counts != counts:
        raise AssertionError("Portable package table counts changed")

    file_sizes = {
        key: path.stat().st_size
        for key, path in artifacts.items()
        if path.exists() and path.is_file()
    }
    report = {
        "status": "PASS",
        "security_code": security_code,
        "first_run_seconds": round(first_seconds, 3),
        "incremental_run_seconds": round(second_seconds, 3),
        "cache_hit": second["cache_hit"],
        "table_counts": counts,
        "parquet_table_count": len(parquet_rows),
        "artifact_file_sizes": file_sizes,
        "portable_copy_status": portable_quality["status"],
        "cross_validation": verify_cross_validation(run_dir),
    }
    report_path = run_dir / "research_pack_w08_e2e_summary.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main(
        Path(sys.argv[1] if len(sys.argv) > 1 else "research-pack-e2e-run"),
        sys.argv[2] if len(sys.argv) > 2 else "002352",
    )
