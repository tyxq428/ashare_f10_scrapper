from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
from openpyxl import load_workbook

from ashare_f10.research_pack.models import ResearchPackArtifacts

EXCEL_CELL_LIMIT = 32_000
RESEARCH_PACK_SCHEMA_VERSION = "1.0.0"


def sha256_file(path: Path | str) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:  # noqa: BLE001
            pass
    return str(value)


def _safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in result.columns:
        if result[column].dtype != "object":
            continue
        result[column] = result[column].map(
            lambda value: (
                json.dumps(value, ensure_ascii=False, separators=(",", ":"))
                if isinstance(value, (dict, list, tuple, set))
                else _json_scalar(value)
            )
        )
        result[column] = result[column].map(
            lambda value: (
                value[: EXCEL_CELL_LIMIT - 40] + "…[完整内容见JSON/Parquet]"
                if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT
                else value
            )
        )
    return result


def _nonempty_frame(frame: pd.DataFrame, fallback_columns: list[str] | None = None) -> pd.DataFrame:
    if frame.shape[1] > 0:
        return frame
    return pd.DataFrame(columns=fallback_columns or ["empty"])


def _write_frame_array(handle: Any, frame: pd.DataFrame, chunk_size: int = 10_000) -> None:
    handle.write("[")
    first = True
    for start in range(0, len(frame), chunk_size):
        payload = frame.iloc[start : start + chunk_size].to_json(
            orient="records",
            force_ascii=False,
            date_format="iso",
        )
        body = payload[1:-1]
        if not body:
            continue
        if not first:
            handle.write(",")
        handle.write(body)
        first = False
    handle.write("]")


def _write_json_package(
    path: Path,
    metadata: dict[str, Any],
    frames: dict[str, pd.DataFrame],
) -> None:
    with path.open("w", encoding="utf-8") as handle:
        handle.write('{"metadata":')
        json.dump(
            metadata,
            handle,
            ensure_ascii=False,
            separators=(",", ":"),
            default=_json_scalar,
        )
        for name, frame in frames.items():
            handle.write(",")
            json.dump(name, handle, ensure_ascii=False)
            handle.write(":")
            _write_frame_array(handle, frame)
        handle.write("}")


def _format_excel(writer: pd.ExcelWriter, sheet_name: str, frame: pd.DataFrame) -> None:
    workbook = writer.book
    sheet = writer.sheets[sheet_name]
    header = workbook.add_format(
        {
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": "#1F4E78",
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "border": 1,
        }
    )
    warning = workbook.add_format({"bg_color": "#FFF2CC"})
    failure = workbook.add_format({"bg_color": "#FCE4D6"})
    success = workbook.add_format({"bg_color": "#E2F0D9"})
    sheet.set_row(0, 28, header)
    sheet.freeze_panes(1, 0)
    if len(frame.columns):
        sheet.autofilter(0, 0, max(len(frame), 1), len(frame.columns) - 1)
    for index, column in enumerate(frame.columns):
        label = str(column).lower()
        if any(token in label for token in ("notes", "source_row", "source_url", "attributes_json")):
            width = 42
        elif any(token in label for token in ("metric", "field", "document", "selection_reason")):
            width = 28
        elif any(token in label for token in ("date", "period", "status", "unit")):
            width = 18
        else:
            width = min(24, max(12, len(str(column)) + 2))
        sheet.set_column(index, index, width)
    status_columns = [
        index for index, column in enumerate(frame.columns) if str(column).lower() in {"status", "source_status"}
    ]
    for index in status_columns:
        if len(frame) == 0:
            continue
        sheet.conditional_format(
            1,
            index,
            len(frame),
            index,
            {"type": "text", "criteria": "containing", "value": "FAIL", "format": failure},
        )
        sheet.conditional_format(
            1,
            index,
            len(frame),
            index,
            {"type": "text", "criteria": "containing", "value": "CONFLICT", "format": failure},
        )
        sheet.conditional_format(
            1,
            index,
            len(frame),
            index,
            {"type": "text", "criteria": "containing", "value": "UNRESOLVED", "format": warning},
        )
        sheet.conditional_format(
            1,
            index,
            len(frame),
            index,
            {"type": "text", "criteria": "containing", "value": "PASS", "format": success},
        )
        sheet.conditional_format(
            1,
            index,
            len(frame),
            index,
            {"type": "text", "criteria": "containing", "value": "VERIFIED", "format": success},
        )


def _write_excel(path: Path, summary: dict[str, Any], frames: dict[str, pd.DataFrame]) -> None:
    sheet_names = {
        "source_facts": "SourceFacts",
        "canonical_observations": "CanonicalFacts",
        "fact_lineage": "FactLineage",
        "evidence_nodes": "EvidenceNodes",
        "evidence_edges": "EvidenceEdges",
        "profit_quality": "ProfitQuality",
        "segments_and_kpis": "SegmentsKPIs",
        "research_and_development": "ResearchRD",
        "capital_structure": "CapitalStructure",
        "capital_events": "CapitalEvents",
        "corporate_governance": "Governance",
        "risk_events": "RiskEvents",
        "coverage_gaps": "CoverageGaps",
        "documents": "Documents",
    }
    with pd.ExcelWriter(
        path,
        engine="xlsxwriter",
        engine_kwargs={"options": {"strings_to_urls": False}},
    ) as writer:
        summary_frame = pd.DataFrame(
            [
                {
                    "metric": key,
                    "value": json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value,
                }
                for key, value in summary.items()
            ]
        )
        summary_frame.to_excel(writer, sheet_name="Summary", index=False)
        _format_excel(writer, "Summary", summary_frame)
        for name, frame in frames.items():
            safe = _safe_frame(_nonempty_frame(frame))
            sheet_name = sheet_names.get(name, name[:31])[:31]
            safe.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_excel(writer, sheet_name, safe)


def _write_duckdb(path: Path, frames: dict[str, pd.DataFrame]) -> None:
    path.unlink(missing_ok=True)
    connection = duckdb.connect(str(path))
    try:
        for name, frame in frames.items():
            table = _nonempty_frame(frame)
            connection.register("incoming", table)
            connection.execute(f'CREATE TABLE "{name}" AS SELECT * FROM incoming')
            connection.unregister("incoming")
        if "canonical_observations" in frames:
            connection.execute(
                """
                CREATE VIEW resolved_canonical_facts AS
                SELECT * FROM canonical_observations
                WHERE status IN ('VERIFIED_MULTI_SOURCE', 'SINGLE_SOURCE')
                """
            )
            connection.execute(
                """
                CREATE VIEW canonical_conflicts AS
                SELECT * FROM canonical_observations
                WHERE status = 'SOURCE_CONFLICT'
                """
            )
            connection.execute(
                "CREATE INDEX idx_canonical_metric ON canonical_observations(metric_id)"
            )
            connection.execute(
                "CREATE INDEX idx_canonical_period ON canonical_observations(report_date, period_type)"
            )
        if "source_facts" in frames:
            connection.execute("CREATE INDEX idx_source_fact_id ON source_facts(source_fact_id)")
            connection.execute("CREATE INDEX idx_source_field ON source_facts(field_key)")
        if "fact_lineage" in frames:
            connection.execute("CREATE INDEX idx_lineage_observation ON fact_lineage(observation_id)")
        if "evidence_edges" in frames:
            connection.execute("CREATE INDEX idx_evidence_from ON evidence_edges(from_node_id)")
            connection.execute("CREATE INDEX idx_evidence_to ON evidence_edges(to_node_id)")
    finally:
        connection.close()


def validate_research_pack(
    artifacts: ResearchPackArtifacts,
    frames: dict[str, pd.DataFrame] | None = None,
) -> dict[str, Any]:
    failures: list[str] = []
    required_files = {
        "manifest": artifacts.manifest_json,
        "summary": artifacts.summary_json,
        "json": artifacts.package_json,
        "excel": artifacts.package_excel,
        "duckdb": artifacts.package_duckdb,
        "checkpoint": artifacts.checkpoint_json,
    }
    for label, path in required_files.items():
        if not path.exists() or path.stat().st_size == 0:
            failures.append(f"missing_or_empty:{label}:{path}")

    if artifacts.package_excel.exists():
        try:
            workbook = load_workbook(artifacts.package_excel, read_only=True, data_only=False)
            if "Summary" not in workbook.sheetnames:
                failures.append("excel_missing_summary_sheet")
            workbook.close()
        except Exception as exc:  # noqa: BLE001
            failures.append(f"excel_open_error:{exc}")

    duckdb_tables: list[str] = []
    if artifacts.package_duckdb.exists():
        try:
            connection = duckdb.connect(str(artifacts.package_duckdb), read_only=True)
            try:
                duckdb_tables = [row[0] for row in connection.execute("SHOW TABLES").fetchall()]
            finally:
                connection.close()
        except Exception as exc:  # noqa: BLE001
            failures.append(f"duckdb_open_error:{exc}")

    counts: dict[str, int] = {}
    if frames is not None:
        counts = {name: len(frame) for name, frame in frames.items()}
        source_facts = frames.get("source_facts", pd.DataFrame())
        canonical = frames.get("canonical_observations", pd.DataFrame())
        lineage = frames.get("fact_lineage", pd.DataFrame())
        nodes = frames.get("evidence_nodes", pd.DataFrame())
        edges = frames.get("evidence_edges", pd.DataFrame())
        if not source_facts.empty and source_facts["source_fact_id"].duplicated().any():
            failures.append("duplicate_source_fact_id")
        if not canonical.empty and canonical["observation_id"].duplicated().any():
            failures.append("duplicate_observation_id")
        if not lineage.empty and lineage["lineage_id"].duplicated().any():
            failures.append("duplicate_lineage_id")
        if not canonical.empty:
            linked = set(lineage.get("observation_id", []))
            missing = canonical[~canonical["observation_id"].isin(linked)]
            if len(missing):
                failures.append(f"observations_without_lineage:{len(missing)}")
        if not nodes.empty and not edges.empty:
            node_ids = set(nodes["node_id"])
            dangling = edges[
                ~edges["from_node_id"].isin(node_ids) | ~edges["to_node_id"].isin(node_ids)
            ]
            if len(dangling):
                failures.append(f"dangling_evidence_edges:{len(dangling)}")

    file_hashes = {
        label: sha256_file(path)
        for label, path in required_files.items()
        if path.exists() and path.stat().st_size > 0
    }
    return {
        "schema_version": RESEARCH_PACK_SCHEMA_VERSION,
        "status": "PASS" if not failures else "FAIL",
        "failures": failures,
        "table_counts": counts,
        "duckdb_tables": duckdb_tables,
        "file_hashes": file_hashes,
    }


class ResearchPackExporter:
    def __init__(self, output_dir: Path | str) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        (self.output_dir / "tables").mkdir(parents=True, exist_ok=True)
        (self.output_dir / "exports").mkdir(parents=True, exist_ok=True)
        (self.output_dir / "quality").mkdir(parents=True, exist_ok=True)

    def write(
        self,
        security_code: str,
        summary: dict[str, Any],
        manifest: dict[str, Any],
        frames: dict[str, pd.DataFrame],
        checkpoint_path: Path,
    ) -> tuple[ResearchPackArtifacts, dict[str, Any]]:
        for name, frame in frames.items():
            _nonempty_frame(frame).to_parquet(self.output_dir / "tables" / f"{name}.parquet", index=False)

        manifest_json = self.output_dir / "manifest.json"
        summary_json = self.output_dir / "summary.json"
        package_json = self.output_dir / "exports" / f"{security_code}_research_pack.json"
        package_excel = self.output_dir / "exports" / f"{security_code}_research_pack.xlsx"
        package_duckdb = self.output_dir / "exports" / f"{security_code}_research_pack.duckdb"
        quality_json = self.output_dir / "quality" / "research_pack_quality.json"

        manifest_json.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2, default=_json_scalar),
            encoding="utf-8",
        )
        summary_json.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2, default=_json_scalar),
            encoding="utf-8",
        )
        _write_json_package(package_json, {**summary, "manifest": manifest}, frames)
        _write_excel(package_excel, summary, frames)
        _write_duckdb(package_duckdb, frames)

        artifacts = ResearchPackArtifacts(
            output_dir=self.output_dir,
            manifest_json=manifest_json,
            summary_json=summary_json,
            package_json=package_json,
            package_excel=package_excel,
            package_duckdb=package_duckdb,
            quality_json=quality_json,
            checkpoint_json=checkpoint_path,
        )
        quality = validate_research_pack(artifacts, frames)
        quality_json.write_text(
            json.dumps(quality, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return artifacts, quality
