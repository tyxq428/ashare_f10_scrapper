from __future__ import annotations

import hashlib
import json
from collections import Counter
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ashare_f10.validation.models import LogicCheck, OfficialDocument, OfficialFact, ReconciliationResult, TTMValidation, ValidationArtifacts


class ValidationReportWriter:
    def __init__(self, output_dir: Path | str) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write(self, security_code: str, documents: list[OfficialDocument], official_facts: list[OfficialFact], reconciliation: list[ReconciliationResult], logic_checks: list[LogicCheck], ttm_checks: list[TTMValidation]) -> ValidationArtifacts:
        summary_path = self.output_dir / "validation_summary.json"
        detail_path = self.output_dir / "validation_detail.parquet"
        official_path = self.output_dir / "official_facts.parquet"
        evidence_path = self.output_dir / "validation_evidence.json"
        excel_path = self.output_dir / "validation_mismatches.xlsx"
        hashes_path = self.output_dir / "source_hashes.json"
        detail_frame = pd.DataFrame([item.to_dict() for item in reconciliation])
        official_frame = pd.DataFrame([item.to_dict() for item in official_facts])
        logic_frame = pd.DataFrame([item.to_dict() for item in logic_checks])
        ttm_frame = pd.DataFrame([item.to_dict() for item in ttm_checks])
        document_frame = pd.DataFrame([item.to_dict() for item in documents])
        detail_frame.to_parquet(detail_path, index=False)
        official_frame.to_parquet(official_path, index=False)
        status_counts = Counter(item.status for item in reconciliation)
        grade_counts = Counter(item.verification_grade for item in reconciliation)
        extracted_by_report = Counter(item.report_date for item in official_facts)
        logic_counts = Counter(item.status for item in logic_checks)
        ttm_counts = Counter(item.status for item in ttm_checks)
        high_severity = [item for item in reconciliation if item.status == "MISMATCH"]
        summary: dict[str, Any] = {"schema_version": "1.0.0", "security_code": security_code, "validation_scope": "FREE_OFFICIAL_SOURCES_THIN_SLICE", "paid_sources_used": False, "official_sources": sorted({item.source for item in documents}), "document_count": len(documents), "official_fact_count": len(official_facts), "reconciliation_count": len(reconciliation), "status_counts": dict(status_counts), "verification_grade_counts": dict(grade_counts), "logic_check_counts": dict(logic_counts), "ttm_check_counts": dict(ttm_counts), "facts_by_report_date": dict(extracted_by_report), "high_severity_mismatch_count": len(high_severity), "manual_review_required": bool(high_severity), "acceptance_status": self._acceptance_status(documents, official_facts, reconciliation, logic_checks, ttm_checks), "artifacts": {"validation_detail_parquet": str(detail_path), "official_facts_parquet": str(official_path), "validation_evidence_json": str(evidence_path), "validation_mismatches_excel": str(excel_path), "source_hashes_json": str(hashes_path)}}
        summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        evidence = {"documents": [item.to_dict() for item in documents], "official_facts": [item.to_dict() for item in official_facts], "reconciliation": [item.to_dict() for item in reconciliation], "logic_checks": [item.to_dict() for item in logic_checks], "ttm_checks": [item.to_dict() for item in ttm_checks]}
        evidence_path.write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
        hashes = {item.local_path: {"source": item.source, "title": item.title, "report_date": item.report_date, "url": item.url, "sha256": item.sha256, "size_bytes": Path(item.local_path).stat().st_size if item.local_path else 0} for item in documents}
        hashes_path.write_text(json.dumps(hashes, ensure_ascii=False, indent=2), encoding="utf-8")
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)
            detail_frame.to_excel(writer, sheet_name="Reconciliation", index=False)
            official_frame.to_excel(writer, sheet_name="OfficialFacts", index=False)
            logic_frame.to_excel(writer, sheet_name="LogicChecks", index=False)
            ttm_frame.to_excel(writer, sheet_name="TTMChecks", index=False)
            document_frame.to_excel(writer, sheet_name="Documents", index=False)
        self._format_excel(excel_path)
        return ValidationArtifacts(self.output_dir, summary_path, detail_path, official_path, evidence_path, excel_path, hashes_path)

    @staticmethod
    def _acceptance_status(documents: list[OfficialDocument], official_facts: list[OfficialFact], reconciliation: list[ReconciliationResult], logic_checks: list[LogicCheck], ttm_checks: list[TTMValidation]) -> str:
        if len(documents) < 2:
            return "FAIL_DOCUMENT_DISCOVERY"
        if not {item.report_date for item in documents}.issubset({item.report_date for item in official_facts}):
            return "FAIL_EXTRACTION"
        if len(official_facts) < 20:
            return "PARTIAL_EXTRACTION"
        if any(item.status == "MISMATCH" for item in reconciliation):
            return "FAIL_SOURCE_CONFLICT"
        if any(item.status == "FAIL" for item in logic_checks):
            return "FAIL_ACCOUNTING_LOGIC"
        if any(item.status == "FAIL" for item in ttm_checks):
            return "FAIL_TTM_CONSISTENCY"
        return "PASS"

    @staticmethod
    def _format_excel(path: Path) -> None:
        workbook = load_workbook(path)
        header_fill, pass_fill, fail_fill = PatternFill("solid", fgColor="1F4E78"), PatternFill("solid", fgColor="E2F0D9"), PatternFill("solid", fgColor="FCE4D6")
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.fill, cell.font = header_fill, Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column in sheet.columns:
                letter = column[0].column_letter
                sheet.column_dimensions[letter].width = max(12, min(60, max((len(str(cell.value or "")) for cell in column), default=10) + 2))
            if sheet.title in {"Reconciliation", "LogicChecks", "TTMChecks"}:
                headers = {cell.value: cell.column for cell in sheet[1]}
                status_column = headers.get("status")
                if status_column:
                    for row in range(2, sheet.max_row + 1):
                        cell = sheet.cell(row=row, column=status_column)
                        cell.fill = pass_fill if cell.value in {"PASS", "EXACT_MATCH", "WITHIN_ROUNDING"} else fail_fill
        workbook.save(path)


def sha256_file(path: Path | str) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
