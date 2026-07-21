from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ashare_f10.raw_sources.models import RawPackArtifacts, RawPackRun, SourceDocument, write_json


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")


def _document_rows(documents: list[SourceDocument]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for document in documents:
        row = document.model_dump(mode="json")
        row["attachment_count"] = len(document.attachments)
        row["attachments"] = json.dumps(row["attachments"], ensure_ascii=False)
        row["links"] = json.dumps(row["links"], ensure_ascii=False)
        row["section_index"] = json.dumps(row["section_index"], ensure_ascii=False)
        row["table_index"] = json.dumps(row["table_index"], ensure_ascii=False)
        rows.append(row)
    return rows


def build_source_index(documents: list[SourceDocument]) -> pd.DataFrame:
    rows = _document_rows(documents)
    return pd.DataFrame(rows)


def _export_excel(documents: list[SourceDocument], quality: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "来源文档索引"
    columns = [
        "document_id",
        "security_code",
        "pack_id",
        "source_id",
        "source_tier",
        "source_organization",
        "status",
        "access_status",
        "document_type",
        "document_title",
        "publish_date",
        "report_period",
        "entity_match_status",
        "entity_match_confidence",
        "source_url",
        "sha256",
        "text_sha256",
        "original_file_path",
        "parsed_text_path",
        "retrieved_at_utc",
        "minimum_human_action",
        "notes",
    ]
    ws.append(columns)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for document in documents:
        payload = document.model_dump(mode="json")
        ws.append([payload.get(column) for column in columns])
    widths = [34, 12, 26, 12, 18, 30, 22, 24, 20, 48, 14, 14, 24, 18, 56, 66, 66, 48, 48, 24, 56, 64]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index) if index <= 26 else "V"].width = width
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    qs = workbook.create_sheet("质量摘要")
    qs.append(["指标", "结果"])
    for cell in qs[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for key, value in quality.items():
        qs.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    qs.column_dimensions["A"].width = 36
    qs.column_dimensions["B"].width = 70
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def validate_raw_pack(output_dir: Path) -> dict[str, Any]:
    manifest_path = output_dir / "metadata" / "raw_pack_run.json"
    documents_path = output_dir / "source_index" / "source_documents.jsonl"
    failures: list[str] = []
    if not manifest_path.exists():
        failures.append("missing run manifest")
    if not documents_path.exists():
        failures.append("missing source_documents.jsonl")
    rows: list[dict[str, Any]] = []
    if documents_path.exists():
        for line in documents_path.read_text(encoding="utf-8").splitlines():
            if line.strip():
                rows.append(json.loads(line))
    ids = [row.get("document_id") for row in rows]
    if len(ids) != len(set(ids)):
        failures.append("duplicate document_id")
    for row in rows:
        if row.get("raw_original_binary_saved"):
            path = Path(row.get("original_file_path") or "")
            if not path.exists() or not row.get("sha256"):
                failures.append(f"missing downloaded source file or hash: {row.get('document_id')}")
        if row.get("status") == "NO_MATCH" and not row.get("query"):
            failures.append(f"NO_MATCH missing query: {row.get('document_id')}")
        if row.get("status") == "PERMISSION_BLOCKED" and not row.get("minimum_human_action"):
            failures.append(f"PERMISSION_BLOCKED missing action: {row.get('document_id')}")
    return {
        "status": "PASS" if not failures else "FAIL",
        "document_count": len(rows),
        "status_counts": dict(Counter(row.get("status") for row in rows)),
        "downloaded_file_count": sum(bool(row.get("raw_original_binary_saved")) for row in rows),
        "parsed_text_count": sum(bool(row.get("parsed_text_saved")) for row in rows),
        "permission_blocked_count": sum(row.get("status") == "PERMISSION_BLOCKED" for row in rows),
        "no_match_count": sum(row.get("status") == "NO_MATCH" for row in rows),
        "failures": failures,
    }


def export_raw_pack(run: RawPackRun, output_dir: Path) -> RawPackArtifacts:
    output_dir.mkdir(parents=True, exist_ok=True)
    metadata_dir = output_dir / "metadata"
    index_dir = output_dir / "source_index"
    quality_dir = output_dir / "quality"
    metadata_dir.mkdir(parents=True, exist_ok=True)
    index_dir.mkdir(parents=True, exist_ok=True)
    quality_dir.mkdir(parents=True, exist_ok=True)

    run_manifest = metadata_dir / "raw_pack_run.json"
    documents_jsonl = index_dir / "source_documents.jsonl"
    attachments_jsonl = index_dir / "attachments.jsonl"
    entity_matches_json = index_dir / "entity_matches.json"
    quality_report_json = quality_dir / "raw_pack_quality.json"
    excel_index = index_dir / "raw_pack_index.xlsx"
    parquet_index = index_dir / "source_documents.parquet"
    duckdb_index = index_dir / "raw_pack.duckdb"

    write_json(run_manifest, run.to_manifest())
    document_rows = [document.model_dump(mode="json") for document in run.documents]
    _write_jsonl(documents_jsonl, document_rows)
    attachments = [
        attachment.model_dump(mode="json")
        for document in run.documents
        for attachment in document.attachments
    ]
    _write_jsonl(attachments_jsonl, attachments)
    write_json(entity_matches_json, [match.model_dump(mode="json") for match in run.entity_matches])

    frame = build_source_index(run.documents)
    parquet_path: str | None = None
    duckdb_path: str | None = None
    try:
        frame.to_parquet(parquet_index, index=False)
        parquet_path = str(parquet_index)
    except Exception:  # noqa: BLE001
        parquet_path = None
    try:
        import duckdb

        connection = duckdb.connect(str(duckdb_index))
        try:
            connection.register("source_documents_frame", frame)
            connection.execute(
                "create or replace table source_documents as select * from source_documents_frame"
            )
        finally:
            connection.close()
        duckdb_path = str(duckdb_index)
    except Exception:  # noqa: BLE001
        duckdb_path = None

    preliminary_quality = {
        "status": "PENDING",
        "document_count": len(run.documents),
        "status_counts": run.status_counts(),
        "error_count": len(run.errors),
        "errors": run.errors,
    }
    _export_excel(run.documents, preliminary_quality, excel_index)
    quality = validate_raw_pack(output_dir)
    quality["run_errors"] = run.errors
    write_json(quality_report_json, quality)
    _export_excel(run.documents, quality, excel_index)

    latest_pointer = output_dir.parent / "latest.json"
    artifacts = RawPackArtifacts(
        output_dir=str(output_dir),
        run_manifest=str(run_manifest),
        source_documents_jsonl=str(documents_jsonl),
        attachments_jsonl=str(attachments_jsonl),
        entity_matches_json=str(entity_matches_json),
        quality_report_json=str(quality_report_json),
        excel_index=str(excel_index),
        parquet_index=parquet_path,
        duckdb_index=duckdb_path,
        latest_pointer=str(latest_pointer),
    )
    write_json(latest_pointer, artifacts.model_dump(mode="json"))
    return artifacts
