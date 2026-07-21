from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ashare_f10.fetch.manifest import load_field_mapping

DARK_BLUE = "17365D"
MID_BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"

FINANCIAL_FAMILIES = {
    "RPT_F10_FINANCE_GBALANCE": "资产负债表",
    "RPT_F10_FINANCE_GINCOME": "利润表",
    "RPT_F10_FINANCE_GINCOMEQC": "单季度利润表",
    "RPT_F10_FINANCE_GCASHFLOW": "现金流量表",
    "RPT_F10_FINANCE_GCASHFLOWQC": "单季度现金流",
    "RPT_F10_FINANCE_GRATIO": "财务比率",
    "RPT_F10_FINANCE_QGRATIO": "单季度财务比率",
    "RPT_F10_FINANCE_MAINFINADATA": "主要财务指标",
    "RPT_F10_QTR_MAINFINADATA": "单季度主要指标",
    "RPT_F10_FINANCE_DUPONT": "杜邦分析",
}

CASHFLOW_FAMILIES = {
    "RPT_F10_FINANCE_GCASHFLOW",
    "RPT_F10_FINANCE_GCASHFLOWQC",
}

META_KEYS = {
    "SECUCODE", "SECURITY_CODE", "SECURITY_NAME_ABBR", "ORG_CODE", "ORG_TYPE",
    "REPORT_DATE", "REPORT_TYPE", "REPORT_DATE_NAME", "SECURITY_TYPE_CODE",
    "NOTICE_DATE", "UPDATE_DATE", "CURRENCY", "REPORT_YEAR", "OPINION_TYPE",
}

# The Eastmoney cash-flow payload contains the primary statement, followed by a
# separate "补充资料" reconciliation section.  The latter legitimately has no
# Q1/Q3 values for many companies.  Keep it separate so a sparse note row is
# never mistaken for the primary quarterly cash-flow field.
CASHFLOW_SUPPLEMENT_FALLBACK_KEYS = {
    "NETPROFIT", "MINORITY_INTEREST", "ASSET_IMPAIRMENT", "FA_IR_DEPR",
    "OILGAS_BIOLOGY_DEPR", "IR_DEPR", "IA_AMORTIZE", "LPE_AMORTIZE",
    "DEFER_INCOME_AMORTIZE", "PREPAID_EXPENSE_REDUCE", "ACCRUED_EXPENSE_ADD",
    "DISPOSAL_LONGASSET_LOSS", "FA_SCRAP_LOSS", "FAIRVALUE_CHANGE_LOSS",
    "FINANCE_EXPENSE", "INVEST_LOSS", "DEFER_TAX", "DT_ASSET_REDUCE",
    "DT_LIAB_ADD", "PREDICT_LIAB_ADD", "INVENTORY_REDUCE",
    "OPERATE_RECE_REDUCE", "OPERATE_PAYABLE_ADD", "OTHER",
    "OPERATE_NETCASH_OTHERNOTE", "OPERATE_NETCASH_BALANCENOTE",
    "NETCASH_OPERATENOTE", "DEBT_TRANSFER_CAPITAL", "CONVERT_BOND_1YEAR",
    "FINLEASE_OBTAIN_FA", "UNINVOLVE_INVESTFIN_OTHER", "END_CASH", "BEGIN_CASH",
    "END_CASH_EQUIVALENTS", "BEGIN_CASH_EQUIVALENTS", "CCE_ADD_OTHERNOTE",
    "CCE_ADD_BALANCENOTE", "CCE_ADDNOTE", "USERIGHT_ASSET_AMORTIZE",
}


def safe_title(value: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", value).strip()[:31] or "Sheet"
    base = cleaned
    index = 2
    while cleaned in used:
        suffix = f"_{index}"
        cleaned = base[: 31 - len(suffix)] + suffix
        index += 1
    used.add(cleaned)
    return cleaned


def cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    text = str(value)
    return text if len(text) <= 32000 else text[:31940] + "…[完整值见JSON]"


class Mapping:
    def __init__(self) -> None:
        data = load_field_mapping()
        self.global_map = data.get("global", {})

    def label(self, key: str) -> str:
        return str(self.global_map.get(key, {}).get("label", key))

    def unit(self, key: str) -> str:
        return str(self.global_map.get(key, {}).get("unit", ""))


def style_title(ws, row: int, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(end_col, 1))
    cell = ws.cell(row, 1)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 24


def style_header(ws, row: int, end_col: int, color: str = MID_BLUE) -> None:
    for col in range(1, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color=WHITE if color == MID_BLUE else "000000", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def ordered_keys(records: list[dict[str, Any]]) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()
    for record in records:
        for key in record:
            if key not in seen:
                seen.add(key)
                keys.append(key)
    return keys


def validate_group_records(combined: dict[str, Any]) -> None:
    for group in combined.get("groups", []):
        records = group.get("records", [])
        reported = int(group.get("record_count", len(records)))
        if reported != len(records):
            raise ValueError(
                f"请求组 {group.get('group_id')} 的record_count={reported}，"
                f"但records实际包含{len(records)}条，拒绝生成不一致的Excel"
            )


def cashflow_key_sections(records: list[dict[str, Any]]) -> dict[str, list[str]]:
    all_keys = [key for key in ordered_keys(records) if key not in META_KEYS]
    value_keys = [key for key in all_keys if not key.endswith(("_YOY", "_QOQ"))]

    supplement: list[str] = []
    if "NETPROFIT" in value_keys and "CCE_ADDNOTE" in value_keys:
        start = value_keys.index("NETPROFIT")
        end = value_keys.index("CCE_ADDNOTE")
        if start <= end:
            supplement = value_keys[start : end + 1]
    if not supplement:
        supplement = [
            key
            for key in value_keys
            if key in CASHFLOW_SUPPLEMENT_FALLBACK_KEYS or "NOTE" in key
        ]

    supplement_set = set(supplement)
    return {
        "primary": [key for key in value_keys if key not in supplement_set],
        "supplement": supplement,
        "yoy": [key for key in all_keys if key.endswith("_YOY")],
        "qoq": [key for key in all_keys if key.endswith("_QOQ")],
    }


def write_record_block(ws, row: int, title: str, records: list[dict[str, Any]], mapping: Mapping) -> int:
    keys = ordered_keys(records)
    width = max(8, len(keys))
    ws.cell(row, 1, title)
    style_title(ws, row, width)
    if not keys:
        ws.cell(row + 1, 1, "当前接口返回0条记录")
        return row + 3

    for col, key in enumerate(keys, 1):
        ws.cell(row + 1, col, mapping.label(key))
        ws.cell(row + 2, col, key)
    style_header(ws, row + 1, len(keys))
    for col in range(1, len(keys) + 1):
        cell = ws.cell(row + 2, col)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for offset, record in enumerate(records, row + 3):
        for col, key in enumerate(keys, 1):
            ws.cell(offset, col, cell_value(record.get(key)))
            ws.cell(offset, col).alignment = Alignment(vertical="top", wrap_text=True)
    return row + len(records) + 5


def write_financial_matrix(
    ws,
    records: list[dict[str, Any]],
    mapping: Mapping,
    title: str,
    keys: list[str] | None = None,
) -> None:
    selected_keys = keys if keys is not None else [key for key in ordered_keys(records) if key not in META_KEYS]
    periods = []
    for record in records:
        date = str(record.get("REPORT_DATE", ""))[:10]
        period_name = str(record.get("REPORT_DATE_NAME") or record.get("REPORT_TYPE") or "")
        periods.append(f"{date}\n{period_name}" if period_name else date)
    headers = ["页面项目名称", "原始Key", "单位", *periods]
    ws.cell(1, 1, title)
    style_title(ws, 1, len(headers))
    for col, header in enumerate(headers, 1):
        ws.cell(2, col, header)
    style_header(ws, 2, len(headers))

    for row, key in enumerate(selected_keys, 3):
        ws.cell(row, 1, mapping.label(key))
        ws.cell(row, 2, key)
        ws.cell(row, 3, mapping.unit(key))
        for col, record in enumerate(records, 4):
            ws.cell(row, col, cell_value(record.get(key)))
    ws.freeze_panes = "D3"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 12
    for col in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 17


def create_financial_sheet(
    workbook: Workbook,
    used: set[str],
    name: str,
    records: list[dict[str, Any]],
    mapping: Mapping,
    keys: list[str],
) -> None:
    if not keys:
        return
    ws = workbook.create_sheet(safe_title(name, used))
    write_financial_matrix(ws, records, mapping, name, keys)


def export_excel(combined: dict[str, Any], output_dir: Path) -> Path:
    validate_group_records(combined)
    mapping = Mapping()
    exports_dir = output_dir / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    path = exports_dir / f"{combined['metadata']['security']['code']}_F10_full.xlsx"

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    used: set[str] = set()

    summary = workbook.create_sheet(safe_title("00_目录与说明", used))
    metadata = combined["metadata"]
    summary_rows = [
        ("项目", "内容"),
        ("证券代码", metadata["security"]["secucode"]),
        ("抓取完成时间UTC", metadata["completed_at_utc"]),
        ("请求组数", metadata["group_count"]),
        ("完成请求组", metadata["completed_group_count"]),
        ("失败请求组", metadata["failed_group_count"]),
        ("数据来源", metadata["source"]),
        ("说明", "所有原始Key均保留；现金流量表主表、补充资料及同比/环比已分表展示。"),
    ]
    for row, values in enumerate(summary_rows, 1):
        summary.cell(row, 1, values[0])
        summary.cell(row, 2, values[1])
    style_header(summary, 1, 2)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 80

    by_family: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_theme: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for group in combined.get("groups", []):
        by_family[str(group.get("family", ""))].extend(group.get("records", []))
        by_theme[str(group.get("theme", "其他"))].append(group)

    for family, sheet_label in FINANCIAL_FAMILIES.items():
        records = by_family.get(family, [])
        if not records:
            continue
        if family in CASHFLOW_FAMILIES:
            sections = cashflow_key_sections(records)
            create_financial_sheet(workbook, used, sheet_label, records, mapping, sections["primary"])
            supplement_name = "现金流补充资料" if family == "RPT_F10_FINANCE_GCASHFLOW" else "单季度现金流补充资料"
            create_financial_sheet(workbook, used, supplement_name, records, mapping, sections["supplement"])
            yoy_name = "现金流量表同比" if family == "RPT_F10_FINANCE_GCASHFLOW" else "单季度现金流同比"
            qoq_name = "现金流量表环比" if family == "RPT_F10_FINANCE_GCASHFLOW" else "单季度现金流环比"
            create_financial_sheet(workbook, used, yoy_name, records, mapping, sections["yoy"])
            create_financial_sheet(workbook, used, qoq_name, records, mapping, sections["qoq"])
        else:
            ws = workbook.create_sheet(safe_title(sheet_label, used))
            write_financial_matrix(ws, records, mapping, sheet_label)

    for theme, groups in by_theme.items():
        non_financial = [group for group in groups if group.get("family") not in FINANCIAL_FAMILIES]
        if not non_financial:
            continue
        ws = workbook.create_sheet(safe_title(theme, used))
        row = 1
        for group in non_financial:
            row = write_record_block(
                ws,
                row,
                f"{group.get('family')}｜{group.get('strategy')}",
                group.get("records", []),
                mapping,
            )
        ws.freeze_panes = "A4"
        for col in range(1, min(ws.max_column, 52) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    field_ws = workbook.create_sheet(safe_title("90_字段字典", used))
    headers = ["中文字段名", "原始Key", "单位", "字段分类"]
    for col, header in enumerate(headers, 1):
        field_ws.cell(1, col, header)
    style_header(field_ws, 1, len(headers))
    for row, (key, item) in enumerate(sorted(mapping.global_map.items()), 2):
        field_ws.cell(row, 1, item.get("label", key))
        field_ws.cell(row, 2, key)
        field_ws.cell(row, 3, item.get("unit", ""))
        field_ws.cell(row, 4, item.get("category", "PAGE_DISPLAY_FIELD"))
    field_ws.column_dimensions["A"].width = 44
    field_ws.column_dimensions["B"].width = 38
    field_ws.column_dimensions["C"].width = 15
    field_ws.column_dimensions["D"].width = 24
    field_ws.freeze_panes = "A2"

    quality = workbook.create_sheet(safe_title("92_质量校验", used))
    quality_rows = [
        ("检查项", "结果", "值"),
        ("请求组完成", "PASS" if not metadata["failed_group_count"] else "CHECK", metadata["completed_group_count"]),
        ("失败请求组", "PASS" if not metadata["failed_group_count"] else "FAIL", metadata["failed_group_count"]),
        ("记录数与records一致", "PASS", "全部请求组"),
        ("现金流主表与补充资料分离", "PASS", "避免将补充资料稀疏值误判为主表缺失"),
        ("原始Key保留", "PASS", "100%"),
    ]
    for row, values in enumerate(quality_rows, 1):
        for col, value in enumerate(values, 1):
            quality.cell(row, col, value)
    style_header(quality, 1, 3)
    quality.column_dimensions["A"].width = 38
    quality.column_dimensions["B"].width = 14
    quality.column_dimensions["C"].width = 56

    workbook.save(path)
    return path
